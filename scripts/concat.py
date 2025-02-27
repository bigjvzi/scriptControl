import os
import requests
import shutil
import string
import openpyxl

def to_unicode(text):
    if not text:
        return ""
    return "".join(f"\\u{ord(char):04x}" for char in text)

def process_excel_and_generate_txt(
    input_file,  # 输入的 .xlsx 文件路径
    output_file,  # 输出的 .txt 文件路径
    col1,  # 第一列的列名或索引 (如 'A' 或 1)
    col2,  # 第二列的列名或索引 (如 'B' 或 2)
    sheet_name,  # 选择的表名
    template,
):

    try:
        # 加载 Excel 文件
        wb = openpyxl.load_workbook(input_file)
        sheet = wb[sheet_name]  # 根据用户选择的表加载

        # 处理列名为字母或索引的情况
        col1_idx = col1 if isinstance(col1, int) else openpyxl.utils.column_index_from_string(col1)
        col2_idx = col2 if isinstance(col2, int) else openpyxl.utils.column_index_from_string(col2)

        encoding_type = len(output_file.split("*"))
        output_file = output_file.split("*")[0]

        # 打开输出文件
        with open(output_file, "w", encoding="utf-8") as txt_file:
            for row in sheet.iter_rows(min_row=2):  # 从第2行开始，跳过表头
                value1 = row[col1_idx - 1].value  # 获取第 col1 列的值
                value2 = row[col2_idx - 1].value  # 获取第 col2 列的值

                # 如果有空值，用空字符串代替
                value1 = value1 if value1 is not None else ""
                value2 = value2 if value2 is not None else ""

                if encoding_type == 2:
                    value2 = to_unicode(str(value2))

                # 将两列的值通过连接符拼接
                combined = template_replace(template,{"A":value1, "B":value2}) + "\n"

                if value1 == "":
                    return

                # 写入到输出文件
                txt_file.write(combined)

        print(f"生成完成！文件已保存到: {output_file}")

    except Exception as e:
        print(f"发生错误: {e}")

def template_replace(template, data):
    """
    根据传入的模板字符串和数据进行替换
    :param template: 模板字符串，使用 $ 标识变量
    :param data: 包含替换数据的字典
    :return: 替换后的字符串
    """
    template_obj = string.Template(template)
    try:
        return template_obj.substitute(data)
    except KeyError as e:
        print(f"Error: 缺少必要的替换数据 {e}")
        return template

def fill_list_with_single_element(lst, length):
    if len(lst) == 1:
        while len(lst) < length:
            lst.append(lst[0])
    return lst

def main(**params):

    selected_file = params["input_folder"]
    wb = openpyxl.load_workbook(selected_file)

    # 列出表名并让用户选择
    sheet_names = wb.sheetnames
    sheet_choice = int(params["tabel_index"]) - 1
    
    selected_sheet = sheet_names[sheet_choice]

    export_file = params["file_name"].split(",") 

    max_length = len(export_file)

    template = params["template"] or "$A = $B"

    for i in range(max_length):
        process_excel_and_generate_txt(
            input_file=selected_file,  # 输入文件名
            sheet_name=selected_sheet,  # 用户选择的表名
            output_file=params["remote_folder"]+ "/" +export_file[i],  # 输出文件名
            col1="A",  # 第一列
            col2=chr(ord("B")+i),  # 第二列
            template=template,
            )

if __name__ == "__main__":
    main()